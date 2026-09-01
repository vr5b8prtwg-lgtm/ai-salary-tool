# -*- coding: utf-8 -*-
"""AI 标注员工薪酬结算网页工具 —— Flask 后端。"""

import csv
import io
import json
import re
import os
import sys
import tempfile

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import parser as P

# 打包后（PyInstaller frozen）：可写数据放在 exe 同目录，模板从内置资源读取
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
    TEMPLATE_DIR = os.path.join(sys._MEIPASS, "templates")
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")

PRICES_FILE = os.path.join(BASE_DIR, "prices.json")
ABNORMAL_FILE = os.path.join(BASE_DIR, "abnormal_amounts.json")

app = Flask(__name__, template_folder=TEMPLATE_DIR)
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024 * 1024  # 最大 2GB
app.config["TEMPLATES_AUTO_RELOAD"] = True  # 修改页面后无需重启

# 单机内存态：解析结果 + 统计结果
STATE = {
    "parsed": None,
    "stats": None,
    "source_name": None,
}




def _default_month(weights):
    """默认选择数据量最大的真实月份（YYYY-MM），同量时取较新者；无真实月份时取任一有数据的分组。"""
    real = {m: w for m, w in weights.items() if re.match(r"^\d{4}-\d{2}$", m)}
    if not real:
        real = dict(weights)
    if not real:
        return None
    return max(real, key=lambda m: (real[m], m))

# ---------- 工具函数 ----------

def _load_prices():
    """读取单价配置 {项目key: 每题单价}。"""
    if not os.path.exists(PRICES_FILE):
        return {}
    try:
        with open(PRICES_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return {str(k): float(v) for k, v in data.items() if isinstance(v, (int, float)) and v >= 0}
    except Exception:
        pass
    return {}


def _save_prices(prices):
    with open(PRICES_FILE, "w", encoding="utf-8") as f:
        json.dump(prices, f, ensure_ascii=False, indent=2)


def _load_abnormal():
    """读取异常表单金额 {员工: {月份: 金额}}"""
    if not os.path.exists(ABNORMAL_FILE):
        return {}
    try:
        with open(ABNORMAL_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {}
        result = {}
        for emp, months in data.items():
            if isinstance(months, dict):
                clean = {str(m): float(v) for m, v in months.items() if isinstance(v, (int, float))}
                if clean:
                    result[str(emp)] = clean
        return result
    except Exception:
        pass
    return {}


def _save_abnormal(amounts):
    with open(ABNORMAL_FILE, "w", encoding="utf-8") as f:
        json.dump(amounts, f, ensure_ascii=False, indent=2)


def _require_stats():
    if STATE["parsed"] is None or STATE["stats"] is None:
        return None
    return STATE["stats"]



def _pre_tax(stats, prices, abnormal, month):
    """按月税前薪酬：当月各项目（题数×单价）合计 + 当月异常金额"""
    rows = []
    missing = []
    grand_total = 0.0
    month_data = stats["monthly"].get(month, {})
    for emp in stats["employees"]:
        subtotal = 0.0
        for p in stats["projects"]:
            key = p["key"]
            qty = month_data.get(emp, {}).get(key, 0.0)
            if qty == 0:
                continue
            price = prices.get(key)
            if price is None:
                if key not in missing:
                    missing.append(key)
            else:
                subtotal += qty * price
        abn = abnormal.get(emp, {}).get(month, 0.0)
        pre_tax = round(subtotal + abn, 2)
        grand_total += pre_tax
        rows.append({
            "employee": emp,
            "subtotal": round(subtotal, 2),
            "abnormal": abn,
            "pre_tax": pre_tax,
        })
    abnormal_month = {emp: abnormal.get(emp, {}).get(month, 0.0) for emp in stats["employees"]}
    return {
        "months": stats["months"],
        "employees": stats["employees"],
        "selected_month": month,
        "rows": rows,
        "missing_prices": missing,
        "grand_total": round(grand_total, 2),
        "abnormal_map": abnormal_month,
    }


def _build_matrix_rows(stats, month, values_func, totals_func):
    """构建矩阵表格行数据。"""
    month_data = stats["monthly"].get(month, {})
    rows = []
    for emp in stats["employees"]:
        cells = [values_func(emp, key, month_data.get(emp, {})) for key in [p["key"] for p in stats["projects"]]]
        rows.append({"employee": emp, "cells": cells, "total": totals_func(emp, month_data.get(emp, {}))})
    col_totals = [
        values_func("__total__", key, month_data)
        for key in [p["key"] for p in stats["projects"]]
    ]
    return rows, col_totals


# ---------- 页面 ----------

@app.route("/")
def index():
    return render_template("index.html")


# ---------- 上传与统计 ----------

@app.route("/api/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if f is None or f.filename == "":
        return jsonify({"error": "请先选择要上传的 Excel 文件"}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in (".xls", ".xlsx", ".xlsm"):
        return jsonify({"error": "仅支持 .xls / .xlsx 文件"}), 400

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    try:
        f.save(tmp.name)
        tmp.close()
        parsed = P.parse_workbook(tmp.name, source_name=f.filename)
    except Exception as e:
        return jsonify({"error": "文件解析失败：%s" % e}), 400
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    stats = P.build_stats(parsed)
    STATE["parsed"] = parsed
    STATE["stats"] = stats
    STATE["source_name"] = f.filename

    return jsonify({"ok": True, "summary": _summary(parsed, stats)})


def _summary(parsed, stats):
    return {
        "source_name": parsed["source_name"],
        "total_qty": round(parsed["total_qty"], 4),
        "employee_count": len(stats["employees"]),
        "project_count": len(stats["projects"]),
        "record_count": len(parsed["records"]),
        "skipped_count": len(parsed["skipped"]),
        "warning_count": len(parsed["warnings"]),
        "months": stats["months"],
    }


@app.route("/api/stats")
def stats():
    s = _require_stats()
    if s is None:
        return jsonify({"error": "请先上传并统计 Excel 文件"}), 400
    parsed = STATE["parsed"]
    month = request.args.get("month")
    if month not in s["months"]:
        weights = {m: sum(sum(e.values()) for e in s["monthly"].get(m, {}).values()) for m in s["months"]}
        month = _default_month(weights)
    rows, col_totals = _build_matrix_rows(
        s, month,
        lambda emp, key, md: round(md.get(key, 0.0), 4) if emp != "__total__" else round(sum(md.get(e, {}).get(key, 0.0) for e in s["employees"]), 4),
        lambda emp, md: round(sum(md.values()), 4) if emp != "__total__" else None,
    )
    return jsonify({
        "summary": _summary(parsed, s),
        "months": s["months"],
        "selected_month": month,
        "employees": s["employees"],
        "projects": s["projects"],
        "rows": rows,
        "col_totals": col_totals,
        "sheet_totals": parsed["sheet_totals"],
        "skipped": parsed["skipped"][:2000],
        "skipped_total": len(parsed["skipped"]),
        "warnings": parsed["warnings"],
        "records_preview": parsed["records"][:500],
        "record_count": len(parsed["records"]),
    })


# ---------- 单价 ----------

@app.route("/api/prices", methods=["GET"])
def get_prices():
    prices = _load_prices()
    s = STATE["stats"]
    projects = []
    if s is not None:
        for p in s["projects"]:
            projects.append({
                "key": p["key"],
                "label": p["label"],
                "task_id": p["task_id"],
                "task_name": p["task_name"],
                "sheet": p["sheet"],
                "price": prices.get(p["key"]),
            })
    else:
        for key, price in prices.items():
            projects.append({"key": key, "label": key, "task_id": "", "task_name": key, "sheet": "", "price": price})
    return jsonify({"projects": projects, "prices": prices})


@app.route("/api/prices", methods=["POST"])
def set_prices():
    data = request.get_json(silent=True) or {}
    prices = data.get("prices")
    if not isinstance(prices, dict):
        return jsonify({"error": "单价数据格式不正确"}), 400
    cleaned = {}
    for k, v in prices.items():
        try:
            fv = float(v)
        except (TypeError, ValueError):
            return jsonify({"error": "单价“%s”不是有效数字" % k}), 400
        if fv < 0:
            return jsonify({"error": "单价“%s”不能为负数" % k}), 400
        cleaned[str(k)] = round(fv, 4)
    _save_prices(cleaned)
    return jsonify({"ok": True, "saved": len(cleaned), "prices": cleaned})


# ---------- 工作量 ----------

@app.route("/api/workload")
def workload():
    parsed = STATE["parsed"]
    if parsed is None:
        return jsonify({"error": "请先上传并统计 Excel 文件"}), 400
    stats = STATE["stats"]
    month = request.args.get("month")
    if month not in stats["months"]:
        month = None
    wl = P.build_workload(parsed, month=month)
    mwl = P.build_workload_monthly(parsed)
    rows = [{"employee": emp, "cells": [wl["matrix"][emp].get(p["key"], 0.0) for p in wl["projects"]], "total": round(sum(wl["matrix"][emp].values()), 4)} for emp in wl["employees"]]
    monthly_rows = [{"employee": emp, "cells": [mwl["matrix"][emp].get(m, 0.0) for m in mwl["months"]], "total": round(sum(mwl["matrix"][emp].values()), 4)} for emp in mwl["employees"]]
    return jsonify({
        "employees": wl["employees"],
        "projects": wl["projects"],
        "rows": rows,
        "col_totals": list(wl["col_totals"].values()),
        "details": wl["details"],
        "total_qty": round(sum(wl["col_totals"].values()), 4),
        "months": stats["months"],
        "selected_month": month,
        "monthly": {
            "months": mwl["months"],
            "rows": monthly_rows,
            "col_totals": list(mwl["col_totals"].values()),
            "details": mwl["details"],
        },
    })


# ---------- 薪酬 ----------



# ---------- 导出 ----------

def _style_header(ws, ncols, title=None):
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        start = 2
    else:
        start = 1
    for c in range(1, ncols + 1):
        cell = ws.cell(row=start, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws.cell(row=start + 1, column=2)
    return start


def _export_stats_xlsx(parsed, stats, month=None):
    wb = Workbook()
    wb.remove(wb.active)

    months = [month] if month else stats["months"]

    # 总览
    ws = wb.create_sheet("总览")
    ws.append(["工作表", "有效记录数", "题数合计", "跳过行数"])
    for name, t in parsed["sheet_totals"].items():
        ws.append([name, t["rows"], t["qty"], t["skipped"]])
    ws.append(["合计", len(parsed["records"]), parsed["total_qty"], len(parsed["skipped"])])
    for c in range(1, 5):
        ws.cell(row=1, column=c).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor="4472C4")

    # 每月矩阵
    for m in months:
        mdata = stats["monthly"].get(m, {})
        header = ["员工"] + [p["label"] for p in stats["projects"]] + ["合计"]
        ws = wb.create_sheet("题数_" + m)
        ws.append(header)
        for emp in stats["employees"]:
            row = [emp] + [round(mdata.get(emp, {}).get(p["key"], 0.0), 4) for p in stats["projects"]] + [round(sum(mdata.get(emp, {}).values()), 4)]
            ws.append(row)
        totals_row = ["合计"] + [round(sum(mdata.get(e, {}).get(p["key"], 0.0) for e in stats["employees"]), 4) for p in stats["projects"]] + [round(sum(sum(mdata.get(e, {}).values()) for e in stats["employees"]), 4)]
        ws.append(totals_row)
        _style_header(ws, len(header), title=m)

    return wb



def _export_workload_monthly_xlsx(parsed):
    mwl = P.build_workload_monthly(parsed)
    wb = Workbook()
    ws = wb.active
    ws.title = "按月汇总"
    header = ["员工"] + list(mwl["months"]) + ["合计"]
    ws.append(header)
    for emp in mwl["employees"]:
        ws.append([emp] + [mwl["matrix"][emp].get(m, 0.0) for m in mwl["months"]] + [round(sum(mwl["matrix"][emp].values()), 4)])
    ws.append(["合计"] + list(mwl["col_totals"].values()) + [round(sum(mwl["col_totals"].values()), 4)])
    _style_header(ws, len(header))

    ws2 = wb.create_sheet("挈月工作量明细")
    h2 = ["员工", "月份", "题数", "原始产值", "记录数", "完成天数", "项目数"]
    ws2.append(h2)
    for d in mwl["details"]:
        ws2.append([d["employee"], d["month"], d["qty"], d["raw"], d["records"], d["days"], d["projects"]])
    _style_header(ws2, len(h2))
    return wb


def _export_workload_xlsx(parsed):
    wl = P.build_workload(parsed)
    wb = Workbook()
    ws = wb.active
    ws.title = "工作量矩阵"
    header = ["员工"] + [p["label"] for p in wl["projects"]] + ["合计"]
    ws.append(header)
    for emp in wl["employees"]:
        row = [emp] + [wl["matrix"][emp].get(p["key"], 0.0) for p in wl["projects"]] + [round(sum(wl["matrix"][emp].values()), 4)]
        ws.append(row)
    total_row = ["合计"] + list(wl["col_totals"].values()) + [round(sum(wl["col_totals"].values()), 4)]
    ws.append(total_row)
    _style_header(ws, len(header))

    ws2 = wb.create_sheet("工作量明细")
    h2 = ["员工", "项目", "任务ID", "来源表", "题数", "原始产值", "记录数", "完成天数", "涉及月份"]
    ws2.append(h2)
    for d in wl["details"]:
        ws2.append([d["employee"], d["project_label"], d["task_id"], d["sheet"], d["qty"], d["raw"], d["records"], d["days"], "、".join(d["months"])])
    _style_header(ws2, len(h2))
    return wb


def _export_detail_xlsx(parsed):
    wb = Workbook()
    ws = wb.active
    ws.title = "明细"
    headers = ["月份", "日期", "来源表", "员工", "账户", "任务名称", "任务ID", "子队列", "原始值", "换算系数", "题数", "备注"]
    ws.append(headers)
    for r in parsed["records"]:
        ws.append([
            r["month"], r["date"], r["sheet"], r["employee"], r["account"],
            r["task_name"], r["task_id"], r["sub_queue"], r["raw_qty"],
            r["factor"], r["qty"], r["remark"],
        ])
    _style_header(ws, len(headers))
    return wb


def _wb_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _csv_response(rows, filename):
    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerows(rows)
    buf = io.BytesIO(sio.getvalue().encode("utf-8-sig"))
    return send_file(
        buf, mimetype="text/csv; charset=utf-8",
        as_attachment=True, download_name=filename,
    )


@app.route("/api/export")
def export():
    parsed = STATE["parsed"]
    stats = STATE["stats"]
    if parsed is None or stats is None:
        return jsonify({"error": "请先上传并统计 Excel 文件"}), 400

    kind = request.args.get("kind", "stats")
    month = request.args.get("month") or None
    fmt = request.args.get("format", "xlsx")
    if month not in stats["months"]:
        month = None
    suffix = fmt if fmt in ("xlsx", "csv") else "xlsx"
    fname = "%s_%s.%s" % (STATE["source_name"] or "结算", "全周期" if not month else month, suffix)
    fname = fname.replace("/", "_").replace("\\", "_")


    if kind == "detail":
        if fmt == "csv":
            headers = ["月份", "日期", "来源表", "员工", "账户", "任务名称", "任务ID", "子队列", "原始值", "换算系数", "题数", "备注"]
            rows = [headers] + [[r["month"], r["date"], r["sheet"], r["employee"], r["account"], r["task_name"],
                                 r["task_id"], r["sub_queue"], r["raw_qty"], r["factor"], r["qty"], r["remark"]]
                                for r in parsed["records"]]
            return _csv_response(rows, fname)
        wb = _export_detail_xlsx(parsed)
        return send_file(_wb_to_bytes(wb), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=fname)

    # stats
    if fmt == "csv":
        months = [month] if month else stats["months"]
        rows = [["月份", "员工", "项目", "任务ID", "来源表", "题数"]]
        for m in months:
            for emp in stats["employees"]:
                for p in stats["projects"]:
                    qty = stats["monthly"][m].get(emp, {}).get(p["key"], 0.0)
                    if qty == 0:
                        continue
                    rows.append([m, emp, p["label"], p["task_id"], p["sheet"], qty])
        return _csv_response(rows, fname)
    wb = _export_stats_xlsx(parsed, stats, month)
    return send_file(_wb_to_bytes(wb), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@app.route("/api/tax")
def tax():
    s = _require_stats()
    if s is None:
        return jsonify({"error": "请先上传并统计 Excel 文件"}), 400
    prices = _load_prices()
    abnormal = _load_abnormal()
    month = request.args.get("month")
    if month not in s["months"]:
        weights = {m: sum(sum(e.values()) for e in s["monthly"].get(m, {}).values()) for m in s["months"]}
        month = _default_month(weights)
    return jsonify(_pre_tax(s, prices, abnormal, month))


@app.route("/api/abnormal_amounts", methods=["POST"])
def set_abnormal():
    data = request.get_json(silent=True) or {}
    month = str(data.get("month") or "")
    amounts = data.get("amounts")
    if not month:
        return jsonify({"error": "缺少月份"}), 400
    if not isinstance(amounts, dict):
        return jsonify({"error": "异常金额数据格式不正确"}), 400
    abnormal = _load_abnormal()
    cleaned = {}
    for k, v in amounts.items():
        try:
            fv = float(v)
        except (TypeError, ValueError):
            return jsonify({"error": "员工“%s”的异常金额不是有效数字" % k}), 400
        if fv < 0:
            return jsonify({"error": "员工“%s”的异常金额不能为负数" % k}), 400
        abnormal.setdefault(str(k), {})[month] = round(fv, 2)
        cleaned[str(k)] = round(fv, 2)
    _save_abnormal(abnormal)
    return jsonify({"ok": True, "month": month, "saved": len(cleaned)})


@app.route("/api/reset", methods=["POST"])
def reset():
    STATE["parsed"] = None
    STATE["stats"] = None
    STATE["source_name"] = None
    return jsonify({"ok": True})


def _find_free_port(host, start=5000, end=5010):
    import socket
    for port in range(start, end + 1):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(0.5)
            # 能连接上说明端口已被占用，连接不上则认为可用
            if s.connect_ex((host, port)) != 0:
                return port
    return start


if __name__ == "__main__":
    # Windows 控制台改用 UTF-8，避免中文乱码
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass

    HOST = "127.0.0.1"
    port = _find_free_port(HOST)
    url = "http://%s:%d" % (HOST, port)

    # 打包后把实际地址写入文件，供启动脚本读取
    if getattr(sys, "frozen", False):
        try:
            with open(os.path.join(BASE_DIR, "启动地址.txt"), "w", encoding="utf-8") as f:
                f.write(url)
        except OSError:
            pass

    print("=" * 50)
    print("AI 标注员工税前薪酬工具已启动")
    print("请在浏览器中打开： %s" % url)
    print("按 Ctrl+C 停止服务")
    print("=" * 50)
    app.run(host=HOST, port=port, debug=False, threaded=True)