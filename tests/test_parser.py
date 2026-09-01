# -*- coding: utf-8 -*-
"""自动校验脚本：可直接运行 python tests/test_parser.py"""

import io
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import parser as P

SAMPLE = r"C:\Users\24620\OneDrive\Desktop\统计.xls"

PASS = 0
FAIL = 0


def check(name, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print("  [通过] " + name)
    else:
        FAIL += 1
        print("  [失败] " + name + ("　" + str(detail) if detail else ""))


def make_sample_xlsx(path):
    wb = openpyxl.Workbook()
    # 表1：直接题数（产值即题数），含空行
    ws = wb.active
    ws.title = "dcc年龄"
    ws.append(["日期", "标注员", "账户名", "任务名称", "任务ID", "子队列", "产值", "备注", ""])
    ws.append(["2026-08-11", "张三", "zs", "任务A", "1001", "q1", 125, ""])
    ws.append(["2026-08-12", "李四", "ls", "任务A", "1001", "q2", 162, ""])
    ws.append([None, None, None, None, None, None, None, None])  # 空行
    ws.append(["2026-08-13", "张三", "zs", "任务B", "1002", "q1", 0, "", "实际是2（漏记）"])  # 无效数量；空表头列补充说明
    ws.append(["2026-08-14", "张三", "", "任务A", "1001", "q3", 10, ""])  # 仅昵称空白
    ws.append(["2027/727", "张三", "zs", "任务A", "1001", "q4", 5, "", "实际是6（填错了）"])  # 月日粘连 -> 2027-07-27；空表头列补充说明
    ws.append(["3027/7/1", "李四", "ls", "任务A", "1001", "q5", 3, ""])  # 年份笔误 -> 2027-07-01
    ws.append(["不是日期", "李四", "ls", "任务A", "1001", "q6", 2, ""])  # 无法识别

    # 表2：按备注换算（每包50题）
    ws2 = wb.create_sheet("动态IP标志性动作采集")
    ws2.append(["日期", "标注员", "账户名", "任务名称", "任务ID", "子队列", "产值", "备注", "一包50题"])
    ws2.append(["2026-08-18", "王五", "ww", "动态IP任务", "2001", "s1", 3, "每个任务里有50题总共150题", ""])
    ws2.append(["2026-08-19", "赵六", "zl", "动态IP任务", "2001", "s2", 2, "每包50题", ""])

    # 表3：无任务ID，按任务名称归组
    ws3 = wb.create_sheet("模型对战")
    ws3.append(["日期", "标注员", "账户名", "任务名称", "任务ID", "子队列", "产值", "截图"])
    ws3.append(["2026-07-16", "王文平", "王文平", "LMArena0716", "", "", 66, ""])
    ws3.append(["2026-07-17", "陈亚铜", "陈亚铜", "图片编辑私有对战", "", "", 250, ""])
    ws3.append(["", "", "", "", "", "", 99999, ""])  # 孤儿合计行

    wb.save(path)


def test_synthetic_xlsx():
    print("== 合成 xlsx 解析 ==")
    path = os.path.join(tempfile.gettempdir(), "synth_test.xlsx")
    make_sample_xlsx(path)
    parsed = P.parse_workbook(path)
    recs = parsed["records"]
    check("有效记录数 = 10", len(recs) == 10, len(recs))
    by_emp = {}
    for r in recs:
        by_emp.setdefault(r["employee"], []).append(r)
    zhang = [r for r in recs if r["employee"] == "张三"]
    check("张三 题数合计 = 140", sum(r["qty"] for r in zhang) == 140)
    wang = [r for r in recs if r["employee"] == "王五"]
    check("王五 换算题数 = 150（3包×50）", len(wang) == 1 and wang[0]["qty"] == 150, wang)
    zhao = [r for r in recs if r["employee"] == "赵六"]
    check("赵六 换算题数 = 100（2包×50）", len(zhao) == 1 and zhao[0]["qty"] == 100, zhao)
    wp = [r for r in recs if r["employee"] == "王文平"]
    check("无任务ID按名称归组：王文平 LMArena0716 66题", len(wp) == 1 and wp[0]["project_key"] == "LMArena0716" and wp[0]["qty"] == 66, wp)
    check("孤儿合计行被跳过", all(r["employee"] != "" for r in recs))
    nick = [r for r in recs if r["account"] == "" and r["date"] == "2026-08-14"]
    check("仅昵称空白按标注员归属（张三 10题）", len(nick) == 1 and nick[0]["employee"] == "张三" and nick[0]["qty"] == 10, nick)
    check("生成昵称空白提示", any(w["type"] == "昵称空白" for w in parsed["warnings"]))
    md27 = [r for r in recs if r["date"] == "2026-07-27"]
    check("月日粘连 2027/727 解析为 2026-07-27（年份归当前年）", len(md27) == 1 and md27[0]["month"] == "2026-07" and md27[0]["qty"] == 5, md27)
    md301 = [r for r in recs if r["date"] == "2026-07-01"]
    check("年份笔误 3027/7/1 修正为 2026-07-01（年份归当前年）", len(md301) == 1 and md301[0]["month"] == "2026-07", md301)
    check("生成日期已修正提示", any(w["type"] == "日期已修正" for w in parsed["warnings"]))
    bad = [r for r in recs if r["date"] is None and r["month"] == "未标注日期" and r["qty"] == 2]
    check("无法识别日期归入未标注日期", len(bad) == 1, bad)
    check("生成日期无法识别提示", any(w["type"] == "日期无法识别" for w in parsed["warnings"]))
    notes = [w for w in parsed["warnings"] if w["type"] == "列外补充说明"]
    check("列外补充说明提示生成", len(notes) >= 1, notes)
    check("列外补充说明含员工/项目/数量/原文",
          any(n.get("employee") == "张三" and n.get("project") == "1001" and n.get("qty") == 5 and "实际是6" in n.get("message", "") for n in notes), notes)
    skip_notes = [x for x in parsed["skipped"] if x.get("note")]
    check("跳过行补充说明被记录", len(skip_notes) >= 1 and "实际是2" in skip_notes[0]["note"], skip_notes)
    check("脏数据行（空行/无效数量/孤儿合计）被跳过", len(parsed["skipped"]) == 3, parsed["skipped"])
    check("月份包含 2026-07 / 2026-08", {"2026-07", "2026-08"}.issubset({r["month"] for r in recs}))
    stats = P.build_stats(parsed)
    check("员工数 = 6", len(stats["employees"]) == 6, stats["employees"])
    check("项目数 = 4", len(stats["projects"]) == 4, [p["key"] for p in stats["projects"]])
    m8 = stats["monthly"]["2026-08"]
    check("2026-08 张三 任务A = 135", m8["张三"]["1001"] == 135, m8.get("张三"))


def test_workload():
    print("== 工作量统计 ==")
    path = os.path.join(tempfile.gettempdir(), "synth_test.xlsx")
    make_sample_xlsx(path)
    parsed = P.parse_workbook(path)
    wl = P.build_workload(parsed)
    check("工作量员工数 = 6", len(wl["employees"]) == 6)
    check("工作量项目数 = 4", len(wl["projects"]) == 4)
    check("全周期总题数 = 873", round(sum(wl["col_totals"].values()), 4) == 873, sum(wl["col_totals"].values()))
    by = {(d["employee"], d["project_key"]): d for d in wl["details"]}
    z = by[("张三", "1001")]
    check("张三 1001：题数 140、记录 3、完成天数 3", z["qty"] == 140 and z["records"] == 3 and z["days"] == 3, z)
    w = by[("王五", "2001")]
    check("王五 2001：题数 150、原始产值 3", w["qty"] == 150 and w["raw"] == 3, w)

    mwl = P.build_workload_monthly(parsed)
    check("按月工作量月份包含 2026-07/2026-08", {"2026-07", "2026-08"}.issubset(mwl["months"]), mwl["months"])
    md = {(d["employee"], d["month"]): d for d in mwl["details"]}
    z = md[("张三", "2026-08")]
    check("张三 2026-08：题数 135、记录 2、天数 2、项目 1", z["qty"] == 135 and z["records"] == 2 and z["days"] == 2 and z["projects"] == 1, z)
    wm = md[("王五", "2026-08")]
    check("王五 2026-08：题数 150、原始产值 3", wm["qty"] == 150 and wm["raw"] == 3, wm)


def test_abnormal_context():
    print("== 异常上下文 ==")
    path = os.path.join(tempfile.gettempdir(), "synth_test.xlsx")
    make_sample_xlsx(path)
    parsed = P.parse_workbook(path)
    fix = [w for w in parsed["warnings"] if w["type"] == "日期已修正"]
    check("日期已修正提示含员工/项目/数量",
          any(w.get("employee") == "李四" and w.get("project") == "1001" and w.get("qty") == 3 for w in fix), fix)
    orphan = [x for x in parsed["skipped"] if x.get("qty") == 99999]
    check("孤儿行跳过记录保留数量 99999", len(orphan) == 1, orphan)


def test_new_category():
    print("== 新大类自动识别 ==")
    import openpyxl as _ox
    path = os.path.join(tempfile.gettempdir(), "newcat_test.xlsx")
    wb = _ox.Workbook()
    ws = wb.active
    ws.title = "dlc"
    ws.append(["日期", "标注员", "账户名", "任务名称", "任务ID", "子队列", "产值", "备注"])
    ws.append(["2026-08-20", "孙七", "sq", "dlc任务1", "3001", "q", 4, "每个任务里有20题总共80题"])
    ws.append(["2026-08-21", "孙七", "sq", "dlc任务2", "3002", "q", 5, ""])
    ws2 = wb.create_sheet("dlc直接题数")
    ws2.append(["日期", "标注员", "账户名", "任务名称", "任务ID", "子队列", "数量"])
    ws2.append(["2026-08-22", "周八", "zb", "直接任务", "4001", "q", 9])
    wb.save(path)

    parsed = P.parse_workbook(path)
    totals = {k: round(v["qty"], 4) for k, v in parsed["sheet_totals"].items()}
    check("新大类 dlc 被识别", "dlc" in totals, totals)
    check("dlc 备注换算：4×20=80 且 5 按产值原值", totals.get("dlc") == 85, totals.get("dlc"))
    check("新大类（数量列）被识别且直接题数", totals.get("dlc直接题数") == 9, totals.get("dlc直接题数"))
    stats = P.build_stats(parsed)
    sheets = {p["sheet"] for p in stats["projects"]}
    check("项目归入新大类", {"dlc", "dlc直接题数"}.issubset(sheets), sheets)
    wl = P.build_workload(parsed)
    wl_sheets = {p["sheet"] for p in wl["projects"]}
    check("工作量按新大类分组", {"dlc", "dlc直接题数"}.issubset(wl_sheets), wl_sheets)


def test_real_sample():
    print("== 真实样例 统计.xls ==")
    if not os.path.exists(SAMPLE):
        print("  [跳过] 样例文件不存在：" + SAMPLE)
        return
    parsed = P.parse_workbook(SAMPLE)
    totals = {k: round(v["qty"], 4) for k, v in parsed["sheet_totals"].items()}
    expect = {
        "Dpo arena 评测": 94800,
        "动态IP标志性动作采集": 1850,
        "dcc年龄": 20285,
        "html评测": 1060,
        "模型对战": 135370,
        "特殊运镜采集": 420,
    }
    for name, exp in expect.items():
        check("表[%s] 题数 = %s" % (name, exp), totals.get(name) == exp, totals.get(name))
    check("总计 = 253785", round(parsed["total_qty"], 4) == 253785, parsed["total_qty"])

    skipped_by_sheet = {}
    for s in parsed["skipped"]:
        skipped_by_sheet[s["sheet"]] = skipped_by_sheet.get(s["sheet"], 0) + 1
    check("Dpo 跳过 7 行", skipped_by_sheet.get("Dpo arena 评测") == 7, skipped_by_sheet.get("Dpo arena 评测"))
    check("模型对战 跳过 374 行", skipped_by_sheet.get("模型对战") == 374, skipped_by_sheet.get("模型对战"))
    check("动态IP 跳过 2 行", skipped_by_sheet.get("动态IP标志性动作采集") == 2, skipped_by_sheet.get("动态IP标志性动作采集"))

    rec = next((r for r in parsed["records"] if r["date"] == "2026-07-17" and r["employee"] == "王文平" and r["task_name"] == "LMArena0716"), None)
    check("抽查：2026-07-17 王文平 LMArena0716 = 66题", rec is not None and rec["qty"] == 66, rec)

    months = sorted({r["month"] for r in parsed["records"]})
    check("月份含 2026-07 与 2026-08", "2026-07" in months and "2026-08" in months, months)

    warns = [w for w in parsed["warnings"] if w["type"] == "日期异常"]
    n420 = [w for w in parsed["warnings"] if w["sheet"] == "Dpo arena 评测" and w.get("row") == 420 and w["type"] == "列外补充说明"]
    n723 = [x for x in parsed["skipped"] if x["sheet"] == "Dpo arena 评测" and x["row"] == 723 and x.get("note")]
    n960 = [x for x in parsed["skipped"] if x["sheet"] == "Dpo arena 评测" and x["row"] == 960 and x.get("note")]
    check("Dpo 723 跳过行补充说明（5）", len(n723) == 1 and "5" in n723[0]["note"], n723)
    check("Dpo 960 跳过行补充说明（18）", len(n960) == 1 and "18" in n960[0]["note"], n960)
    check("Dpo 第420行 列外补充说明（含 22）", len(n420) >= 1 and "22" in n420[0].get("message", ""), n420)
    check("存在日期异常提示", len(warns) >= 1, len(warns))


def test_app_endpoints():
    print("== 后端接口（Flask 测试客户端）==")
    import app as A
    A.STATE["parsed"] = None
    A.STATE["stats"] = None
    # 清理残留配置，保证可重复运行
    for _f in (A.PRICES_FILE, A.ABNORMAL_FILE):
        if os.path.exists(_f):
            os.remove(_f)
    client = A.app.test_client()

    path = os.path.join(tempfile.gettempdir(), "synth_test.xlsx")
    with open(path, "rb") as f:
        data = {"file": (io.BytesIO(f.read()), "synth_test.xlsx")}
        r = client.post("/api/upload", data=data, content_type="multipart/form-data")
    check("上传接口 200", r.status_code == 200, r.status_code)
    js = r.get_json()
    check("上传后总题数 = 873", js["summary"]["total_qty"] == 873, js["summary"].get("total_qty"))

    r = client.get("/api/stats")
    check("统计接口 200", r.status_code == 200)
    stats = r.get_json()
    check("统计含 6 名员工", len(stats["employees"]) == 6, len(stats["employees"]))

    r = client.post("/api/prices", json={"prices": {"1001": 1.5, "1002": 2.0, "LMArena0716": 0.8}})
    check("设置单价 200", r.status_code == 200)

    r = client.get("/api/export?kind=stats&format=csv")
    check("导出统计 CSV 200", r.status_code == 200)
    r = client.get("/api/export?kind=detail&format=csv")
    check("导出明细 CSV 200", r.status_code == 200)

    r = client.get("/api/workload")
    wl = r.get_json()
    check("工作量接口 200 且总题数 873", r.status_code == 200 and abs(wl["total_qty"] - 873) < 0.01, wl.get("total_qty"))
    r = client.get("/api/export?kind=workload&format=csv")
    check("导出工作量 CSV 200", r.status_code == 200)
    r = client.get("/api/export?kind=workload&format=xlsx")
    check("导出工作量 Excel 200", r.status_code == 200 and r.data[:2] == b"PK", len(r.data))

    r = client.get("/api/workload")
    wl2 = r.get_json()
    check("工作量接口含按月数据", "monthly" in wl2 and len(wl2["monthly"]["months"]) >= 2, wl2.get("monthly", {}).get("months"))
    r = client.get("/api/workload?month=2026-08")
    wlm = r.get_json()
    check("工作量接口按月筛选 2026-08 合计 547",
          r.status_code == 200 and abs(wlm["total_qty"] - 547) < 0.01 and wlm["selected_month"] == "2026-08", (wlm.get("total_qty"), wlm.get("selected_month")))
    r = client.get("/api/export?kind=workload_monthly&format=csv")
    check("导出按月工作量 CSV 200", r.status_code == 200)
    r = client.get("/api/export?kind=workload_monthly&format=xlsx")
    check("导出挈月工作量 Excel 200", r.status_code == 200 and r.data[:2] == b"PK", len(r.data))

    # 税前薪酬按月：价格 1001=1.5、LMArena0716=0.8；默认月份为题数最多的 2026-08
    r = client.get("/api/tax")
    tx = r.get_json()
    check("税前薪酬接口 200 且默认月 2026-08 合计 445.5",
          r.status_code == 200 and tx.get("selected_month") == "2026-08" and abs(tx["grand_total"] - 445.5) < 0.01, (tx.get("selected_month"), tx.get("grand_total")))
    zhang = next((x for x in tx["rows"] if x["employee"] == "张三"), None)
    check("张三 2026-08 项目合计 = 202.5", zhang is not None and abs(zhang["subtotal"] - 202.5) < 0.01, zhang)

    r = client.post("/api/abnormal_amounts", json={"month": "2026-08", "amounts": {"张三": 100.0}})
    check("保存异常金额 200", r.status_code == 200)
    r = client.get("/api/tax?month=2026-08")
    tx2 = r.get_json()
    zhang2 = next((x for x in tx2["rows"] if x["employee"] == "张三"), None)
    check("张三 2026-08 应发 = 302.5（含异常 100），合计 545.5",
          zhang2 is not None and abs(zhang2["pre_tax"] - 302.5) < 0.01 and abs(tx2["grand_total"] - 545.5) < 0.01, (zhang2, tx2.get("grand_total")))
    r = client.get("/api/tax?month=2026-07")
    tx3 = r.get_json()
    check("2026-07 税前合计 = 64.8（LMArena0716 66×0.8 + 1001 8×1.5）",
          abs(tx3["grand_total"] - 64.8) < 0.01, tx3.get("grand_total"))


if __name__ == "__main__":
    test_synthetic_xlsx()
    test_workload()
    test_abnormal_context()
    test_new_category()
    test_real_sample()
    test_app_endpoints()
    print("\n结果：通过 %d 项，失败 %d 项" % (PASS, FAIL))
    sys.exit(1 if FAIL else 0)