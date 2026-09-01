# -*- coding: utf-8 -*-
"""Excel 解析模块：读取 .xls / .xlsx，按规则换算标注题数。"""

import os
import re
from collections import OrderedDict, defaultdict
from datetime import datetime, timedelta

import xlrd
import openpyxl

# 需要从“备注”中提取每包/每任务题数的工作表
REMARK_FACTOR_SHEETS = {"动态IP标志性动作采集", "特殊运镜采集"}

# 表头关键词（按优先级排列）
HEADER_RULES = [
    ("date", ["日期"]),
    ("employee", ["标注员", "人员", "姓名"]),
    ("account", ["账户名"]),
    ("task_name", ["任务名称", "任务名"]),
    ("task_id", ["任务ID", "任务id", "任务编号", "任务 Id"]),
    ("sub_queue", ["子队列", "子列队"]),
    ("qty", ["实际题数", "产值", "题数", "数量"]),
    ("remark", ["备注"]),
]


def _norm(v):
    """浮点整数值转 int，便于展示。"""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def _to_text(v):
    """任意单元格值转干净文本。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_number(v):
    """尝试把单元格值转成数值，失败返回 None。"""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("，", "")
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _excel_col_name(idx):
    """将 0 基列索引转换为 Excel 列名（A, B, ..., Z, AA...）。"""
    idx += 1
    name = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        name = chr(65 + rem) + name
    return name


def _excel_serial_to_datetime(serial):
    return datetime(1899, 12, 30) + timedelta(days=float(serial))


def _parse_text_date(s):
    s = s.strip().replace("年", "/").replace("月", "/").replace("日", "")
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # 月日粘连：2027/727 -> 2027/7/27；2027/1127 -> 2027/11/27
    m = re.match(r"^(\d{4})[/.\-](\d{3})$", s)
    if m:
        y = int(m.group(1)); mon = int(m.group(2)[0]); day = int(m.group(2)[1:])
        if 1 <= mon <= 9 and 1 <= day <= 31:
            try:
                return datetime(y, mon, day)
            except ValueError:
                pass
    m = re.match(r"^(\d{4})[/.\-](\d{4})$", s)
    if m:
        y = int(m.group(1)); mon = int(m.group(2)[:2]); day = int(m.group(2)[2:])
        if 1 <= mon <= 12 and 1 <= day <= 31:
            try:
                return datetime(y, mon, day)
            except ValueError:
                pass
    try:
        return _excel_serial_to_datetime(float(s))
    except ValueError:
        return None


def _cell_datetime(value, kind, datemode=None):
    """把单元格转成 datetime；无法识别返回 None。"""
    if kind == "date":
        if datemode is not None and isinstance(value, (int, float)):
            return xlrd.xldate_as_datetime(value, datemode)
        if isinstance(value, datetime):
            return value
        return None
    if kind == "number":
        return _excel_serial_to_datetime(value)
    if kind == "text":
        return _parse_text_date(value)
    return None


def _read_xls(path):
    """读取 .xls：逐表返回 (表名, 表头, 行列表)。行内单元格为 (值, 类型)。"""
    wb = xlrd.open_workbook(path, on_demand=True)
    datemode = wb.datemode
    for sheet_name in wb.sheet_names():
        sh = wb.sheet_by_name(sheet_name)
        headers = []
        if sh.nrows > 0:
            headers = [sh.cell_value(0, c) for c in range(sh.ncols)]
        rows = []
        for r in range(1, sh.nrows):
            cells = []
            for c in range(sh.ncols):
                v = sh.cell_value(r, c)
                t = sh.cell_type(r, c)
                if t == 0 or t == 6:
                    cells.append((None, "empty"))
                elif t == 3:
                    cells.append((v, "date"))
                elif t == 1:
                    cells.append((v, "text"))
                elif t == 2:
                    cells.append((v, "number"))
                else:
                    cells.append((v, "text"))
            rows.append((r, cells))
        yield sheet_name, headers, rows, datemode


def _read_xlsx(path):
    """读取 .xlsx：逐表返回 (表名, 表头, 行列表)。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        rows = []
        row_idx = 0
        for r, row in enumerate(ws.iter_rows(), start=1):
            if r == 1:
                headers = [c.value for c in row]
                continue
            cells = []
            for c in row:
                v = c.value
                if v is None:
                    cells.append((None, "empty"))
                elif isinstance(v, datetime):
                    cells.append((v, "date"))
                elif isinstance(v, (int, float)):
                    cells.append((v, "number"))
                else:
                    cells.append((str(v), "text"))
            rows.append((r, cells))
        yield sheet_name, headers, rows, None


def read_workbook(path):
    """按扩展名选择读取器。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        yield from _read_xls(path)
    elif ext == ".xlsx" or ext == ".xlsm":
        yield from _read_xlsx(path)
    else:
        raise ValueError("不支持的文件格式：%s（请上传 .xls 或 .xlsx）" % ext)


def detect_columns(headers):
    """根据表头识别各列索引。返回 {字段: 索引}。"""
    result = {}
    header_texts = [_to_text(h) for h in headers]
    used = set()
    for field, keywords in HEADER_RULES:
        for i, h in enumerate(header_texts):
            if i in used or not h:
                continue
            for kw in keywords:
                if kw in h:
                    result[field] = i
                    used.add(i)
                    break
            if field in result and result[field] == i:
                break
    return result


def _extract_factor(remark):
    """从备注中提取“X题”，返回整数；无则返回 None。"""
    if not remark:
        return None
    m = re.search(r"(\d+)\s*题", remark)
    if m:
        return int(m.group(1))
    return None


def parse_workbook(path, source_name=None):
    """解析整个工作簿，返回统计所需全部数据。"""
    records = []
    skipped = []
    warnings = []
    sheet_totals = OrderedDict()
    current_year = datetime.now().year

    for sheet_name, headers, rows, datemode in read_workbook(path):
        cols = detect_columns(headers)
        qty_idx = cols.get("qty")
        emp_idx = cols.get("employee")
        acc_idx = cols.get("account")
        tn_idx = cols.get("task_name")
        tid_idx = cols.get("task_id")
        sq_idx = cols.get("sub_queue")
        rm_idx = cols.get("remark")
        dt_idx = cols.get("date")

        if qty_idx is None or (emp_idx is None and acc_idx is None) or (tn_idx is None and tid_idx is None):
            warnings.append({
                "sheet": sheet_name,
                "type": "表结构无法识别",
                "message": "该工作表缺少必要的表头（日期/标注员/任务名称/任务ID/产值等），已整表跳过",
            })
            sheet_totals[sheet_name] = {"rows": 0, "qty": 0.0, "skipped": 0}
            continue

        qty_header = _to_text(headers[qty_idx]) if qty_idx < len(headers) else ""
        is_direct = "实际题数" in qty_header
        use_remark_factor = (not is_direct) and (sheet_name in REMARK_FACTOR_SHEETS)

        # 表外补充列：表头为空的列，员工可能在锁表后在这里补写说明
        extra_cols = [c for c, h in enumerate(headers) if not _to_text(h).strip()]

        def extra_note_text(cells):
            parts = []
            for c in extra_cols:
                if c >= len(cells):
                    continue
                xv = cells[c][0]
                if xv is None:
                    continue
                xt = _to_text(xv)
                if xt:
                    parts.append("第%s列「%s」" % (_excel_col_name(c), xt[:120]))
            return "；".join(parts)

        sheet_qty = 0.0
        sheet_rows = 0
        sheet_skipped = 0
        nick_blank = 0

        for row_no, cells in rows:
            def cell(idx):
                if idx is None or idx >= len(cells):
                    return (None, "empty")
                return cells[idx]

            raw_all = [_norm(v) for v, _ in cells]
            if all(v in (None, "") for v in raw_all):
                sheet_skipped += 1
                skipped.append({"sheet": sheet_name, "row": row_no + 1, "reason": "空行"})
                continue

            ev, _ = cell(emp_idx)
            av, _ = cell(acc_idx)
            tv, _ = cell(tn_idx)
            idv, _ = cell(tid_idx)
            employee = _to_text(ev)
            account = _to_text(av)
            task_name = _to_text(tv)
            task_id = _to_text(idv)
            project_key = task_id or task_name

            # 先解析日期（含年份笔误修正），得到归属月份
            dvv, dvk = cell(dt_idx)
            dt_raw = _cell_datetime(dvv, dvk, datemode)
            month = "未标注日期"
            date_iso = None
            corrected = None
            if dt_raw is not None:
                if dt_raw.year > current_year:
                    try:
                        corrected = dt_raw.replace(year=current_year)
                    except ValueError:
                        corrected = None
                eff = corrected if corrected is not None else dt_raw
                month = eff.strftime("%Y-%m")
                date_iso = eff.strftime("%Y-%m-%d")
            orig_iso = dt_raw.strftime("%Y-%m-%d") if dt_raw is not None else None

            qv, qk = cell(qty_idx)
            qnum = _to_number(qv)
            if qnum is None or qnum <= 0:
                sheet_skipped += 1
                _n = extra_note_text(cells)
                skipped.append({
                    "sheet": sheet_name, "row": row_no + 1, "reason": "数量无效（无有效题数/产值）",
                    "employee": employee or account, "project": project_key, "qty": None, "month": month, "date": date_iso, "orig_date": orig_iso if corrected is not None else None,
                    **({"note": _n} if _n else {}),
                })
                continue

            if not employee and not account:
                sheet_skipped += 1
                _n = extra_note_text(cells)
                skipped.append({
                    "sheet": sheet_name, "row": row_no + 1, "reason": "缺少员工和账户",
                    "project": project_key, "qty": qnum, "month": month, "date": date_iso, "orig_date": orig_iso if corrected is not None else None,
                    **({"note": _n} if _n else {}),
                })
                continue

            if not project_key:
                sheet_skipped += 1
                _n = extra_note_text(cells)
                skipped.append({
                    "sheet": sheet_name, "row": row_no + 1, "reason": "缺少项目（任务ID/任务名称均为空）",
                    "employee": employee or account, "qty": qnum, "month": month, "date": date_iso, "orig_date": orig_iso if corrected is not None else None,
                    **({"note": _n} if _n else {}),
                })
                continue

            sv, _ = cell(sq_idx)
            rv, _ = cell(rm_idx)
            remark = _to_text(rv)

            employee_final = employee or account
            label = task_name or project_key

            factor = 1
            qty = qnum
            if not is_direct:
                f = _extract_factor(remark)
                if f:
                    factor = f
                    qty = qnum * f
                elif use_remark_factor:
                    warnings.append({
                        "sheet": sheet_name,
                        "row": row_no + 1,
                        "employee": employee_final,
                        "project": project_key,
                        "qty": qnum,
                        "month": month,
                        "date": date_iso,
                        "orig_date": orig_iso if corrected is not None else None,
                        "type": "换算系数缺失",
                        "message": "备注中未找到“X题”的换算系数，按产值原值统计",
                    })

            # 日期提示（按修正后的归属月份）
            if corrected is not None:
                warnings.append({
                    "sheet": sheet_name,
                    "row": row_no + 1,
                    "employee": employee_final,
                    "project": project_key,
                    "qty": qty,
                    "month": month,
                    "date": date_iso,
                    "orig_date": orig_iso if corrected is not None else None,
                    "type": "日期已修正",
                    "message": "日期 %s 疑似年份笔误（超出当前年份），已按 %s 修正（仅调整年份，月日不变）" % (dt_raw.strftime("%Y-%m-%d"), corrected.strftime("%Y-%m-%d")),
                })
            elif dt_raw is not None and (dt_raw.year < current_year - 2 or dt_raw.year > current_year):
                warnings.append({
                    "sheet": sheet_name,
                    "row": row_no + 1,
                    "employee": employee_final,
                    "project": project_key,
                    "qty": qty,
                    "month": month,
                    "date": date_iso,
                    "orig_date": orig_iso if corrected is not None else None,
                    "type": "日期异常",
                    "message": "日期 %s 可能录入有误，已归入 %s 月份" % (date_iso, month),
                })
            elif dt_raw is None:
                dv_raw = _to_text(dvv)
                if dv_raw:
                    warnings.append({
                        "sheet": sheet_name,
                        "row": row_no + 1,
                        "employee": employee_final,
                        "project": project_key,
                        "qty": qty,
                        "month": month,
                        "date": date_iso,
                        "orig_date": orig_iso if corrected is not None else None,
                        "type": "日期无法识别",
                        "message": "日期“%s”格式无法识别，已归入“未标注日期”统计" % dv_raw,
                    })
                else:
                    warnings.append({
                        "sheet": sheet_name,
                        "row": row_no + 1,
                        "employee": employee_final,
                        "project": project_key,
                        "qty": qty,
                        "month": month,
                        "date": date_iso,
                        "orig_date": orig_iso if corrected is not None else None,
                        "type": "日期缺失",
                        "message": "该行未填写日期，已归入“未标注日期”统计",
                    })

            for c in extra_cols:
                xv, _ = cell(c)
                if xv is None:
                    continue
                xt = _to_text(xv)
                if not xt:
                    continue
                warnings.append({
                    "sheet": sheet_name,
                    "row": row_no + 1,
                    "employee": employee_final,
                    "project": project_key,
                    "qty": qty,
                    "month": month,
                    "date": date_iso,
                    "orig_date": orig_iso if corrected is not None else None,
                    "type": "列外补充说明",
                    "message": "第 %s 列（表头为空）有补充内容：「%s」" % (_excel_col_name(c), xt[:120]),
                })

            if employee and not account:
                nick_blank += 1

            records.append({
                "sheet": sheet_name,
                "row": row_no + 1,
                "date": date_iso,
                "month": month,
                "employee": employee_final,
                "account": account,
                "task_name": task_name,
                "task_id": task_id,
                "project_key": project_key,
                "project_label": label,
                "sub_queue": _to_text(sv),
                "raw_qty": qnum,
                "factor": factor,
                "qty": qty,
                "remark": remark,
            })
            sheet_qty += qty
            sheet_rows += 1

        sheet_totals[sheet_name] = {"rows": sheet_rows, "qty": sheet_qty, "skipped": sheet_skipped}
        if nick_blank:
            warnings.append({
                "sheet": sheet_name,
                "type": "昵称空白",
                "message": "%d 行昵称（账户名）空白，已按该行标注员归属" % nick_blank,
            })

    total_qty = sum(r["qty"] for r in records)
    return {
        "source_name": source_name or os.path.basename(path),
        "records": records,
        "skipped": skipped,
        "warnings": warnings,
        "sheet_totals": sheet_totals,
        "total_qty": total_qty,
    }



def build_workload(parsed, month=None):
    """工作量：每个员工×每个项目的题数、原始产值、记录数、完成天数；month 为 None 时为全周期。"""
    all_records = parsed["records"]
    records = [r for r in all_records if r["month"] == month] if month else all_records

    # 项目、员工始终用全量，保证下拉选项完整
    projects = OrderedDict()
    for r in all_records:
        projects.setdefault(r["project_key"], {
            "key": r["project_key"],
            "label": r["project_label"],
            "task_id": r["task_id"],
            "task_name": r["task_name"],
            "sheet": r["sheet"],
        })

    agg = defaultdict(lambda: {"qty": 0.0, "raw": 0.0, "records": 0, "days": set(), "months": set()})
    for r in records:
        d = agg[(r["employee"], r["project_key"])]
        d["qty"] += r["qty"]
        d["raw"] += r["raw_qty"]
        d["records"] += 1
        if r["date"]:
            d["days"].add(r["date"])
        d["months"].add(r["month"])

    employees = sorted({r["employee"] for r in all_records})
    matrix = {}
    col_totals = {}
    for emp in employees:
        row = {}
        for pj in projects:
            key = pj
            val = agg.get((emp, key), {}).get("qty", 0.0)
            row[key] = round(val, 4)
        matrix[emp] = row
    for key in projects:
        col_totals[key] = round(sum(agg.get((emp, key), {}).get("qty", 0.0) for emp in employees), 4)

    details = []
    for (emp, key), d in agg.items():
        pj = projects[key]
        details.append({
            "employee": emp,
            "project_key": key,
            "project_label": pj["label"],
            "task_id": pj["task_id"],
            "task_name": pj["task_name"],
            "sheet": pj["sheet"],
            "qty": round(d["qty"], 4),
            "raw": round(d["raw"], 4),
            "records": d["records"],
            "days": len(d["days"]),
            "months": sorted(d["months"]),
        })
    details.sort(key=lambda x: (x["employee"], x["project_label"]))

    projects_list = list(projects.values())
    return {
        "employees": employees,
        "projects": projects_list,
        "matrix": matrix,
        "col_totals": col_totals,
        "details": details,
    }


def build_workload_monthly(parsed):
    """按月工作量：每个员工在每个月份的题数、原始产值、记录数、完成天数、项目数。"""
    records = parsed["records"]
    raw_months = sorted({r["month"] for r in records})
    months = [m for m in raw_months if re.match(r"^\d{4}-\d{2}$", m)] + [m for m in raw_months if not re.match(r"^\d{4}-\d{2}$", m)]

    agg = defaultdict(lambda: {"qty": 0.0, "raw": 0.0, "records": 0, "days": set(), "projects": set()})
    for r in records:
        d = agg[(r["employee"], r["month"])]
        d["qty"] += r["qty"]
        d["raw"] += r["raw_qty"]
        d["records"] += 1
        if r["date"]:
            d["days"].add(r["date"])
        d["projects"].add(r["project_key"])

    employees = sorted({r["employee"] for r in records})
    matrix = {}
    for emp in employees:
        matrix[emp] = {m: round(agg.get((emp, m), {}).get("qty", 0.0), 4) for m in months}
    col_totals = {m: round(sum(agg.get((emp, m), {}).get("qty", 0.0) for emp in employees), 4) for m in months}

    details = []
    for (emp, m), d in agg.items():
        details.append({
            "employee": emp,
            "month": m,
            "qty": round(d["qty"], 4),
            "raw": round(d["raw"], 4),
            "records": d["records"],
            "days": len(d["days"]),
            "projects": len(d["projects"]),
        })
    details.sort(key=lambda x: (x["employee"], x["month"]))

    return {
        "months": months,
        "employees": employees,
        "matrix": matrix,
        "col_totals": col_totals,
        "details": details,
    }


def build_stats(parsed):
    """基于解析结果构建统计：月度、员工、项目、题数。"""
    records = parsed["records"]
    raw_months = sorted({r["month"] for r in records})
    months = [m for m in raw_months if re.match(r"^\d{4}-\d{2}$", m)] + [m for m in raw_months if not re.match(r"^\d{4}-\d{2}$", m)]

    # project 元数据：key -> {label, task_id, task_name, sheet}
    projects = OrderedDict()
    for r in records:
        p = projects.setdefault(r["project_key"], {
            "key": r["project_key"],
            "label": r["project_label"],
            "task_id": r["task_id"],
            "task_name": r["task_name"],
            "sheet": r["sheet"],
        })

    # month -> employee -> project -> qty
    monthly = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in records:
        monthly[r["month"]][r["employee"]][r["project_key"]] += r["qty"]

    employees = sorted({r["employee"] for r in records})

    projects_list = [
        {
            "key": p["key"],
            "label": p["label"],
            "task_id": p["task_id"],
            "task_name": p["task_name"],
            "sheet": p["sheet"],
        }
        for p in projects.values()
    ]
    return {
        "months": months,
        "employees": employees,
        "projects": projects_list,
        "monthly": {
            m: {
                emp: {k: round(v, 4) for k, v in projs.items()}
                for emp, projs in emps.items()
            }
            for m, emps in monthly.items()
        },
    }